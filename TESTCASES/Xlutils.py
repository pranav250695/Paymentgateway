import openpyxl


def GetRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)


def GetColCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)


def ReadData(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=columnno).value


def WriteData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=columnno).value = data
    workbook.save(file)